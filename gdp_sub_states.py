from openpyxl import Workbook
import json
from openpyxl.chart import (
    ScatterChart,
    Series,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active
colors=['FF0000',"FF9933", "00FFCC", "FFFFCC",  "CC99FF", "FF66FF", "99FF99", "CC3399","003366", "CCCCFF", "9900FF", "660066", "990099", "FF0066", "FF9933", "333300", "66330", "669999"]
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=14)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=14)
main_title="Services"
states="""[
             {
                "state":"Khartoum",
                "years":[
                   {
                      "sub":[
                         {
                            "title":"electrical",
                            "value":"2000"
                         },
                         {
                            "title":"transportation",
                            "value":"5007"
                         },
                         {
                            "title":"civil",
                            "value":"7000"
                         }
                      ],
                      "year":"2013"
                   },
                   {
                      "sub":[
                         {
                            "title":"electrical",
                            "value":"4746"
                         },
                         {
                            "title":"transportation",
                            "value":"7654"
                         },
                         {
                            "title":"civil",
                            "value":"3455"
                         }
                      ],
                      "year":"2014"
                   },
                   {
                      "sub":[
                         {
                            "title":"electrical",
                            "value":"2564"
                         },
                         {
                            "title":"transportation",
                            "value":"567"
                         },
                         {
                            "title":"civil",
                            "value":"3456"
                         }
                      ],
                      "year":"2015"
                   }
                ]
             },
             {
                "state":"Northern",
                "years":[
                   {
                      "sub":[
                         {
                            "title":"electrical",
                            "value":"8000"
                         },
                         {
                            "title":"transportation",
                            "value":"9000"
                         },
                         {
                            "title":"civil",
                            "value":"8500"
                         }
                      ],
                      "year":"2013"
                   },
                   {
                      "sub":[
                         {
                            "title":"electrical",
                            "value":"4664"
                         },
                         {
                            "title":"transportation",
                            "value":"4465"
                         },
                         {
                            "title":"civil",
                            "value":"3455"
                         }
                      ],
                      "year":"2014"
                   },
                   {
                      "sub":[
                         {
                            "title":"electrical",
                            "value":"3534"
                         },
                         {
                            "title":"transportation",
                            "value":"235"
                         },
                         {
                            "title":"civil",
                            "value":"2443"
                         }
                      ],
                      "year":"2015"
                   }
                ]
             }
          ]"""
states=json.loads(states)
rowIndex=1
for state in states:

    ws.merge_cells('A{}:B{}'.format(rowIndex,rowIndex))
    ws.cell(column=1, row=rowIndex, value='{} State'.format(state['state']))
    ws['A{}'.format(rowIndex)].style = highlight
    rowIndex+=1
    years=state['years']
    startRow=rowIndex
    for year in years:
        year['sub'].sort()
    if len(years)>0:
        ws.cell(column=1, row=rowIndex, value='Year')
        col=2
        for sub in years[0]["sub"]:
           ws.cell(column=col, row=rowIndex, value=sub["title"])
           col+=1
        rowIndex+=1
    for year in years:
        ws.cell(column=1, row=rowIndex, value=year["year"])
        col=2
        for sub in year["sub"]:
            ws.cell(column=col, row=rowIndex, value=float(sub["value"]))
            col+=1
        rowIndex+=1
    #data = Reference(ws, min_col=2, min_row=startRow, max_col=len(states[0]['years'][0]["sub"])+startRow, max_row=rowIndex-1)
    c2 = ScatterChart()
    c2.title = "GDP of "+main_title+" at "+state["state"]+" State"
    c2.style = 12
    c2.y_axis.title = "Value"
    c2.x_axis.title = "Year"

    #c2.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=startRow+1, max_row=rowIndex)
    #c2.set_categories(dates)
    colorIndex=0
    for i in range(2,2+len(states[0]["years"][0]["sub"])):
        values = Reference(ws, min_col=i, min_row=startRow, max_row=rowIndex-1)
        series = Series(values, dates, title_from_data=True)
        series.smooth = True
        series.graphicalProperties.line.width = 50000
        series.graphicalProperties.line.solidFill = colors[colorIndex]
        c2.series.append(series)
        colorIndex+=1
        if colorIndex>=len(colors):
            colorIndex=0
    ws.add_chart(c2, "A{}".format(rowIndex))
    rowIndex+=16


wb.save("gdp_sub_states.xlsx")
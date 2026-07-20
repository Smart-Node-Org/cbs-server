from openpyxl import Workbook
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
import json
import sys

from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

years=sys.argv[1]
file_name=sys.argv[2]

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=15)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=13)
years=json.loads(years)
cellIndex=1
yearIndex=0
shift=0
startGraph=0
for year in years:
    year['rates'].sort()
    ws.merge_cells('A{}:B{}'.format(cellIndex,cellIndex))
    ws.cell(column=1, row=cellIndex, value='Year {}'.format(year['year']))
    ws['A{}'.format(cellIndex)].style = highlight
    cellIndex+=1
    ws.cell(column=1, row=cellIndex, value='Sector')
    ws.cell(column=2, row=cellIndex, value='Contribution')
    ws['A{}'.format(cellIndex)].style = bold
    ws['B{}'.format(cellIndex)].style = bold
    cellIndex+=1
    startGraph=cellIndex
    for rate in year["rates"]:
        ws.cell(column=1, row=cellIndex, value=rate['sector'])
        ws.cell(column=2, row=cellIndex, value=rate['Contribution_rate'])
        cellIndex+=1
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=startGraph, max_row=cellIndex)
    data = Reference(ws, min_col=2, min_row=startGraph-1, max_row=cellIndex)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Sectors Contribution in {}".format(year['year'])
    pie.width=7
    ws.add_chart(pie, "D{}".format(startGraph-2))
    if len(year["rates"])>14:
        cellIndex+=len(year["rates"])+2
    else:
        cellIndex=startGraph+12

dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
print(dims)
ws.column_dimensions['A'].width = dims[1]
ws.column_dimensions['B'].width = dims[2]+5

wb.save(file_name+".xlsx")
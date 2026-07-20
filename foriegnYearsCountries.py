from openpyxl import Workbook
import json
from openpyxl.chart import (
    BarChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=14)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=14)
countries='[{"years": [{"year": "2040", "export": 112, "import": 22, "re_export": 12}, {"year": "2020", "export": 1223, "import": 976, "re_export": 323}], "country": "Afghanistan"}, {"years": [{"year": "2020", "export": 5778, "import": 6756, "re_export": 688}], "country": "Algeria"}]'
countries=json.loads(countries)
rowIndex=1

for country in countries:
    ws.merge_cells('A{}:C{}'.format(rowIndex,rowIndex))
    ws.cell(column=1, row=rowIndex, value=country['country'])
    ws['A{}'.format(rowIndex)].style = highlight
    rowIndex+=1
    years=country['years']
    startRow=rowIndex
    ws.cell(row=rowIndex, column=1, value='Year')
    ws.cell(row=rowIndex, column=2, value='Export')
    ws.cell(row=rowIndex, column=3, value='Re-Export')
    ws.cell(row=rowIndex, column=4, value='Import')
    ws.cell(row=rowIndex, column=5, value='Trade Balance')
    rowIndex+=1
    for year in years:
        ws.cell(column=1, row=rowIndex, value=year['year'])
        ws.cell(column=2, row=rowIndex, value=year['export'])
        ws.cell(column=3, row=rowIndex, value=year['re_export'])
        ws.cell(column=4, row=rowIndex, value=year['import'])
        ws.cell(column=5, row=rowIndex, value=year['re_export']+year['export']-year['import'])
        rowIndex+=1

    data = Reference(ws, min_col=2, min_row=startRow, max_row=len(country['years'])+startRow, max_col=5)
    cats = Reference(ws, min_col=1, min_row=startRow+1, max_row=len(country['years'])+startRow)


    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Foreign Trade between Sudan and {}".format(country['country'])
    chart1.y_axis.title = 'SDG'
    chart1.x_axis.title = 'Years'
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "A{}".format(rowIndex+1))
    rowIndex+=16
wb.save("foriegn_countries.xlsx")
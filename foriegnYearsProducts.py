from openpyxl import Workbook
import json
from openpyxl.chart import (
    BarChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=14)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=14)
years='[{"year": "2019", "products": [{"product": "rice", "value_SDG": 45, "value_unit": 16, "export_import": 1}, {"product": "simsim", "value_SDG": 22, "value_unit": 13, "export_import": 0}, {"product": "wheat", "value_SDG": 0, "value_unit": 0, "export_import": 1}]}, {"year": "2020", "products": [{"product": "rice", "value_SDG": 22, "value_unit": 33, "export_import": 1}, {"product": "wheat", "value_SDG": 2, "value_unit": 1, "export_import": 1}, {"product": "simsim", "value_SDG": 22, "value_unit": 56, "export_import": 0}]}, {"year": "2021", "products": [{"product": "simsim", "value_SDG": 56, "value_unit": 12, "export_import": 0}, {"product": "rice", "value_SDG": 0, "value_unit": 0, "export_import": 1}, {"product": "wheat", "value_SDG": 0, "value_unit": 0, "export_import": 1}]}]'
years=json.loads(years)
rowIndex=1
ws.merge_cells('A{}:E{}'.format(rowIndex,rowIndex))
ws.cell(column=1, row=rowIndex, value="Foreign Trade By Products")
ws['A{}'.format(rowIndex)].style = highlight
rowIndex+=1
ws.cell(row=rowIndex, column=1, value='Year')
col=2
for year in years:
    year['products'].sort()
for product in years[0]['products']:
    ws.cell(row=rowIndex, column=col, value=product['product'])
    col+=1
rowIndex+=1
startRow=rowIndex
for year in years:
    ws.cell(row=rowIndex, column=1, value=year['year'])
    col=2
    for product in year['products']:
        ws.cell(column=col, row=rowIndex, value=product['value_SDG'])
        col+=1
    rowIndex+=1
data = Reference(ws, min_col=2, min_row=startRow-1, max_row=rowIndex, max_col=len(years[0]['products'])+1)
cats = Reference(ws, min_col=1, min_row=startRow, max_row=rowIndex)


chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Foreign Trade By Products"
chart1.y_axis.title = 'SDG'
chart1.x_axis.title = 'Years'
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "A{}".format(rowIndex+1))
wb.save("foriegn_products.xlsx")
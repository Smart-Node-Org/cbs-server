import openpyxl
import json
import sys
path = sys.argv[1]

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
years = []
for i in range(2,sheet_obj.max_column,2):
    years.append(int(sheet_obj.cell(row = 1, column = i).value))
products = []
for i in range(3,sheet_obj.max_row+1):
    products.append({"product":str(sheet_obj.cell(row = i, column = 1).value),"years":[]})
    for j in range(len(years)):
        products[len(products)-1]["years"].append({
            "year":years[j],
            "unit":int(sheet_obj.cell(row = len(products)+2, column = j*2+2).value),
            "value":int(sheet_obj.cell(row = len(products)+2, column = j*2+3).value)
            })
print(json.dumps(products))

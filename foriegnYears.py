from openpyxl import Workbook
import json
from openpyxl.chart import (
    BarChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=14)
years='[{"year": "1987", "export": 22, "import": 20, "re_export": 22}, {"year": "1989", "export": 121, "import": 22, "re_export": 22}, {"year": "1990", "export": 65, "import": 96, "re_export": 56}]'
years=json.loads(years)
rowIndex=1
ws.cell(row=1, column=1, value='Year')
ws.cell(row=1, column=2, value='Export')
ws.cell(row=1, column=3, value='Re-Export')
ws.cell(row=1, column=4, value='Import')
ws.cell(row=1, column=5, value='Trade Balance')
rowIndex=2
for year in years:
    ws.cell(column=1, row=rowIndex, value=year['year'])
    ws.cell(column=2, row=rowIndex, value=year['export'])
    ws.cell(column=3, row=rowIndex, value=year['re_export'])
    ws.cell(column=4, row=rowIndex, value=year['import'])
    ws.cell(column=5, row=rowIndex, value=year['re_export']+year['export']-year['import'])
    rowIndex+=1

data = Reference(ws, min_col=2, min_row=1, max_row=len(years), max_col=5)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(years))


chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Foreign Trade"
chart1.y_axis.title = 'SDG'
chart1.x_axis.title = 'Years'
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "A{}".format(rowIndex+1))
wb.save("foriegn.xlsx")
import openpyxl
import json
import sys
path = "Population.xlsx"#sys.argv[1]

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
groups = []
for i in range(2,sheet_obj.max_column,4):
    groups.append(sheet_obj.cell(row = 1, column = i).value.replace('"',''))
states = []
for i in range(4,sheet_obj.max_row+1):
    states.append({"state":str(sheet_obj.cell(row = i, column = 1).value),"population":[]})
    for j in range(len(groups)):
        states[len(states)-1]["population"].append({
            "group":groups[j],
            "rural_male_value":int(sheet_obj.cell(row = len(states)+3, column = j*4+2).value),
            "rural_female_value":int(sheet_obj.cell(row = len(states)+3, column = j*4+3).value),
            "urban_male_value":int(sheet_obj.cell(row = len(states)+3, column = j*4+4).value),
            "urban_female_value":int(sheet_obj.cell(row = len(states)+3, column = j*4+5).value)
            })
print(json.dumps(states))

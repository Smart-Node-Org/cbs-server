from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    ProjectedPieChart,
    Reference
)

import json
import sys

from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

states=sys.argv[1]
file_name=sys.argv[2]
charType=sys.argv[3]
selectedYear=sys.argv[4]

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=15)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=13)
# [{state:"kassala",data:[{year:2012,rural:2323,urban:234},{year:2012,rural:2323,urban:234}]}]
states=json.loads(states)
states=[state for state in states if len(state['data'])>0]

yearIndex=0
shift=0
startGraph=0
rowIndex=1
for state in states:
    print rowIndex
    years=state['data']
    ws.merge_cells('A{}:D{}'.format(rowIndex,rowIndex))
    ws.cell(row=rowIndex, column=1, value="CPI "+charType[0].upper()+charType[1:]+" of "+state['state'])
    ws['A{}'.format(rowIndex)].style = highlight
    rowIndex+=1
    startRow=rowIndex
    ws.cell(row=rowIndex, column=1, value='Month')
    col=2
    for key in years[0]:
        if key != 'year' and key != 'Month':
            ws.cell(row=rowIndex, column=col, value=key)
            col+=1
    rowIndex+=1

    for year in years:
        ws.cell(column=1, row=rowIndex, value=year['Month'])
        col=2
        for key in year:
            if key != 'year' and key != 'Month':
                ws.cell(row=rowIndex, column=col, value=year[key])
                col+=1
        rowIndex+=1


    data = Reference(ws, min_col=2, min_row=startRow, max_row=startRow+len(years)+1, max_col=len(years[0]))
    cats = Reference(ws, min_col=1, min_row=startRow+1, max_row=startRow+len(years)+1)



    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "CPI "+charType[0].upper()+charType[1:]+" of "+state['state']
    chart1.y_axis.title = 'Months'
    chart1.x_axis.title = 'Values'
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "A{}".format(rowIndex+1))
    rowIndex+=17
wb.save(file_name+".xlsx")
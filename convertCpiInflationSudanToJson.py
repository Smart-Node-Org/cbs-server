import openpyxl
import json
import sys
path = sys.argv[1]

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
years = []
for i in range(3,sheet_obj.max_row):
    years.append(int(sheet_obj.cell(row = i, column = 1).value))
rows = []
for i in range(3,sheet_obj.max_row+1):
    rows.append({"year":str(sheet_obj.cell(row = i, column = 1).value),"urban_standard":str(sheet_obj.cell(row = i, column = 2).value),"urban_ongoing":str(sheet_obj.cell(row = i, column = 3).value),"rural_standard":str(sheet_obj.cell(row = i, column = 4).value),"rural_ongoing":str(sheet_obj.cell(row = i, column = 5).value)})

print(json.dumps(rows))

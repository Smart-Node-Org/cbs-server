from openpyxl import Workbook
import json
from openpyxl.chart import (
    LineChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=13)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=14)
years='[{"year": "2015", "title": "services sector", "tot_gdp": 46656}, {"year": "2016", "title": "services sector", "tot_gdp": 7777}, {"year": "2017", "title": "services sector", "tot_gdp": 567}, {"year": "2018", "title": "services sector", "tot_gdp": 98387}]'
years=json.loads(years)
cellIndex=1
yearIndex=0
ws.merge_cells('A{}:D{}'.format(cellIndex,cellIndex))
ws.cell(column=1, row=cellIndex, value='GDP Department {}'.format(years[0]['title']))
ws['A{}'.format(cellIndex)].style = highlight
cellIndex+=1
ws.cell(column=1, row=cellIndex, value='Year')
ws.cell(column=2, row=cellIndex, value='Total GDP')
cellIndex+=1
for year in years:
    ws.cell(column=1, row=cellIndex, value=year['year'])
    ws.cell(column=2, row=cellIndex, value=year['tot_gdp'])
    cellIndex+=1

data = Reference(ws, min_col=2, min_row=2, max_col=len(years)+1, max_row=cellIndex-1)


c2 = LineChart()
c2.title = "Total GDP of {}".format(years[0]['title'])
c2.style = 12
c2.y_axis.title = "Value"
c2.x_axis.title = "Year"

c2.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=3, max_row=cellIndex-1)
c2.set_categories(dates)

ws.add_chart(c2, "A{}".format(cellIndex+1))

wb.save("gdpMain.xlsx")

from openpyxl import Workbook
import json
from openpyxl.chart import (
    LineChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import NamedStyle, Font, Border, Side
wb = Workbook()
ws = wb.active

years=sys.argv[1]
file_name=sys.argv[2]

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
#bd = Side(style='thick', color="000000")
#highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
bold=NamedStyle(name="bold")
bold.font = Font(bold=True, size=14)

years=json.loads(years)

products=[2,3]
rowIndex=1
product_index=0

for year in years:
    year['products'].sort()

    for product1 in year['products']:
        p_duplication=0
        p_index=0
        for product2 in year['products']:
            if product1['id']==product2['id']:
                p_duplication+=1
                duplication_index=p_index
            p_index+=1
        if p_duplication>1:
            year["products"].pop(duplication_index)
        if not product1["main_value"]:
            product1["main_value"]=0
        if product1['id'] not in products:
            year['products'].remove(product1)

if len(years)>0:
    ws.cell(column=1, row=1, value='Year')
    col=2
    for product in years[0]["products"]:
       ws.cell(column=col, row=1, value=product["product"])
       col+=1
rowIndex+=1
for year in years:
    ws.cell(column=1, row=rowIndex, value=year["year"])
    col=2
    for product in year["products"]:
        ws.cell(column=col, row=rowIndex, value=product["main_value"])
        col+=1
    rowIndex+=1
data = Reference(ws, min_col=2, min_row=1, max_col=len(years[0]["products"])+1, max_row=rowIndex-1)


c2 = LineChart()
c2.title = "Sectorial Products"
c2.style = 12
c2.y_axis.title = "Value"
c2.x_axis.title = "Year"

c2.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=rowIndex-1)
c2.set_categories(dates)

ws.add_chart(c2, "A{}".format(rowIndex+1))

wb.save(file_name+".xlsx")